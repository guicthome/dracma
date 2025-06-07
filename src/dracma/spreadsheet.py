"""Utilities for loading Excel spreadsheets used by the Dracma agent."""
from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, List, Optional

from openpyxl import load_workbook


@dataclass
class SheetData:
    name: str
    headers: List[str]
    rows: List[Dict[str, Optional[str]]]


def load_sheet(workbook_path: str, sheet_name: str) -> SheetData:
    """Load a sheet from an Excel workbook returning header and rows."""
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb[sheet_name]

    headers: List[str] = []
    rows: List[Dict[str, Optional[str]]] = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        values = list(row)
        if i == 0:
            headers = [str(v) if v is not None else "" for v in values]
            continue
        record = {headers[j]: values[j] for j in range(len(headers))}
        rows.append(record)
    return SheetData(name=sheet_name, headers=headers, rows=rows)
